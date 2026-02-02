"""
Reportes View - Vista del centro de reportes con todos los tipos disponibles.
Integra completamente con DialogoReportes y ReportingEngine existentes.
"""
from __future__ import annotations
from typing import Optional
from datetime import datetime
import os

from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QFrame, QPushButton, QScrollArea,
    QGroupBox, QMessageBox, QFileDialog
)
from PyQt6.QtGui import QColor, QPalette, QGuiApplication

from app.core.db_adapter import DatabaseAdapter

# ✅ CORRECCIÓN: Importar reportes existentes con manejo de errores mejorado
DIALOGO_REPORTES_AVAILABLE = False
REPORTING_ENGINE_AVAILABLE = False
REPORT_WINDOW_AVAILABLE = False

try:
    from app.ui.dialogs.dialogo_reportes import DialogoReportes
    DIALOGO_REPORTES_AVAILABLE = True
    print("[INFO] ✓ DialogoReportes importado correctamente")
except ImportError as e:
    print(f"[WARNING] ✗ DialogoReportes no disponible: {e}")
except Exception as e:
    print(f"[ERROR] ✗ Error importando DialogoReportes: {e}")

try:
    from app.core.reporting import ReportingEngine
    REPORTING_ENGINE_AVAILABLE = True
    print("[INFO] ✓ ReportingEngine importado correctamente")
except ImportError as e:
    print(f"[WARNING] ✗ ReportingEngine no disponible: {e}")
except Exception as e:
    print(f"[ERROR] ✗ Error importando ReportingEngine: {e}")

try:
    from app.ui.windows.reporte_window import ReportWindow
    REPORT_WINDOW_AVAILABLE = True
    print("[INFO] ✓ ReportWindow importado correctamente")
except ImportError as e:
    print(f"[WARNING] ✗ ReportWindow no disponible: {e}")
except Exception as e:
    print(f"[ERROR] ✗ Error importando ReportWindow: {e}")


class ReportCard(QFrame):
    """Card interactiva para un tipo de reporte."""
    
    clicked = pyqtSignal()
    
    def __init__(self, title: str, description: str, icon: str, color: str, parent=None):
        super().__init__(parent)
        self.title = title
        self.description = description
        
        self.setStyleSheet(f"""
            QFrame {{
                background-color: #2D2D30;
                border: 2px solid #3E3E42;
                border-radius: 12px;
                padding: 20px;
            }}
            QFrame:hover {{
                border-color: {color};
                background-color: #37373D;
            }}
        """)
        
        self.setMinimumHeight(180)
        self.setMinimumWidth(220)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        
        # Icono
        icon_label = QLabel(icon)
        icon_label.setStyleSheet(f"""
            QLabel {{
                font-size: 48pt;
                color: {color};
            }}
        """)
        icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(icon_label)
        
        # Título
        title_label = QLabel(title)
        title_label.setStyleSheet(f"""
            QLabel {{
                font-size: 13pt;
                font-weight: bold;
                color: {color};
            }}
        """)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label.setWordWrap(True)
        layout.addWidget(title_label)
        
        # Descripción
        desc_label = QLabel(description)
        desc_label.setStyleSheet("""
            QLabel {
                font-size: 9pt;
                color: #B9C0CC;
            }
        """)
        desc_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        desc_label.setWordWrap(True)
        layout.addWidget(desc_label)
        
        layout.addStretch()
        
        # Botón
        btn = QPushButton("Generar")
        btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px;
                font-weight: bold;
                font-size: 10pt;
            }}
            QPushButton:hover {{
                background-color: {self._darken_color(color)};
            }}
            QPushButton:pressed {{
                background-color: {self._darken_color(color, 0.3)};
            }}
        """)
        btn.clicked.connect(self.clicked.emit)
        layout.addWidget(btn)
    
    def _darken_color(self, hex_color: str, factor: float = 0.15) -> str:
        """Oscurece un color hex."""
        color = QColor(hex_color)
        h, s, v, a = color.getHsv()
        v = max(0, int(v * (1 - factor)))
        color.setHsv(h, s, v, a)
        return color.name()
    
    def mousePressEvent(self, event):
        """Emite señal al hacer clic en la card."""
        if event.button() == Qt.MouseButton.LeftButton:
            self.clicked.emit()
        super().mousePressEvent(event)


class ReportesView(QWidget):
    """
    Vista del centro de reportes integrado con sistemas existentes.
    """
    
    def __init__(self, db: Optional[DatabaseAdapter] = None, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.db = db
        
        # Inicializar ReportingEngine si está disponible
        self.reporting = None
        if REPORTING_ENGINE_AVAILABLE and db:
            try:
                self.reporting = ReportingEngine(db)
            except Exception as e:
                print(f"[WARNING] No se pudo inicializar ReportingEngine: {e}")
        
        # Obtener colores del tema
        self._resolve_theme_colors()
        
        # Configurar UI
        self._setup_ui()
    
    def _resolve_theme_colors(self):
        """Obtiene colores del tema activo."""
        app = QGuiApplication.instance()
        pal: QPalette = app.palette() if app else QPalette()
        
        def get_color(role: QPalette.ColorRole, fallback: str) -> str:
            try:
                color = pal.color(role)
                if color.isValid():
                    return color.name()
            except Exception:
                pass
            return fallback
        
        self.colors = {
            "accent": get_color(QPalette.ColorRole.Highlight, "#7C4DFF"),
            "text": get_color(QPalette.ColorRole.Text, "#E6E9EF"),
            "text_sec": get_color(QPalette.ColorRole.PlaceholderText, "#B9C0CC"),
            "window": get_color(QPalette.ColorRole.Window, "#1E1E1E"),
            "base": get_color(QPalette.ColorRole.Base, "#252526"),
            "border": get_color(QPalette.ColorRole.Mid, "#3A4152"),
            "success": "#00C853",
            "danger": "#FF5252",
            "warning": "#FFA726",
            "info": "#448AFF",
        }
    
    def _setup_ui(self):
        """Configura la interfaz."""
        # Layout principal con scroll
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(20)
        
        # Scroll area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        
        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.setSpacing(25)
        
        # Título
        title = QLabel("📊 Centro de Reportes")
        title.setStyleSheet(f"""
            QLabel {{
                font-size: 28pt;
                font-weight: bold;
                color: {self.colors['accent']};
                padding: 10px 0;
            }}
        """)
        content_layout.addWidget(title)
        
        # Subtítulo
        subtitle = QLabel("Genera reportes detallados y analiza KPIs de tus licitaciones")
        subtitle.setStyleSheet(f"""
            QLabel {{
                font-size: 12pt;
                color: {self.colors['text_sec']};
                margin-bottom: 10px;
            }}
        """)
        content_layout.addWidget(subtitle)
        
        # Grid de reportes
        reports_grid = QGridLayout()
        reports_grid.setSpacing(20)
        
        # ✅ Definir reportes disponibles
        reportes = [
            {
                "title": "Reporte Individual",
                "description": "Análisis completo de una licitación específica con gráficos",
                "icon": "📈",
                "color": self.colors["accent"],
                "action": self._generar_individual,
                "enabled": REPORT_WINDOW_AVAILABLE,
            },
            {
                "title": "KPIs y Métricas",
                "description": "Dashboard completo de indicadores clave con filtros",
                "icon": "🎯",
                "color": self.colors["info"],
                "action": self._abrir_dialogo_reportes,
                "enabled": DIALOGO_REPORTES_AVAILABLE,
            },
            {
                "title": "Reporte Mensual",
                "description": "Genera reporte del mes actual con estadísticas",
                "icon": "📅",
                "color": "#4CAF50",
                "action": self._generar_mensual,
                "enabled": REPORTING_ENGINE_AVAILABLE,
            },
            {
                "title": "Exportar a Excel",
                "description": "Exporta todas las licitaciones con KPIs a Excel",
                "icon": "📗",
                "color": "#00BCD4",
                "action": self._exportar_excel,
                "enabled": REPORTING_ENGINE_AVAILABLE,
            },
            {
                "title": "Análisis Financiero",
                "description": "Reporte detallado de montos, ofertas y efectividad",
                "icon": "💰",
                "color": "#FF9800",
                "action": self._analisis_financiero,
                "enabled": True,
            },
            {
                "title": "Licitaciones Ganadas",
                "description": "Listado de todas las licitaciones adjudicadas",
                "icon": "🏆",
                "color": "#FFC107",
                "action": self._reporte_ganadas,
                "enabled": True,
            },
        ]
        
        # Crear cards
        for i, reporte in enumerate(reportes):
            row = i // 3
            col = i % 3
            
            card = ReportCard(
                title=reporte["title"],
                description=reporte["description"],
                icon=reporte["icon"],
                color=reporte["color"]
            )
            
            if reporte["enabled"]:
                card.clicked.connect(reporte["action"])
            else:
                card.setEnabled(False)
                card.setToolTip("Módulo no disponible")
            
            reports_grid.addWidget(card, row, col)
        
        content_layout.addLayout(reports_grid)
        
        # Sección de acciones rápidas
        quick_section = self._create_quick_actions_section()
        content_layout.addWidget(quick_section)
        
        content_layout.addStretch()
        
        scroll.setWidget(content)
        main_layout.addWidget(scroll)
    
    def _create_quick_actions_section(self) -> QGroupBox:
        """Crea sección de acciones rápidas."""
        box = QGroupBox("⚡ Acciones Rápidas")
        box.setStyleSheet(f"""
            QGroupBox {{
                background-color: {self.colors['base']};
                border: 1px solid {self.colors['border']};
                border-radius: 12px;
                padding: 20px;
                font-size: 12pt;
                font-weight: bold;
                color: {self.colors['accent']};
                margin-top: 10px;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0 8px;
            }}
        """)
        
        layout = QHBoxLayout(box)
        layout.setSpacing(15)
        
        # Botón: Abrir carpeta de reportes
        btn_abrir_carpeta = QPushButton("📁 Abrir Carpeta de Reportes")
        btn_abrir_carpeta.clicked.connect(self._abrir_carpeta_reportes)
        btn_abrir_carpeta.setStyleSheet(f"""
            QPushButton {{
                background-color: {self.colors['info']};
                color: white;
                border: none;
                border-radius: 8px;
                padding: 12px 20px;
                font-size: 11pt;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: #3367D6;
            }}
        """)
        layout.addWidget(btn_abrir_carpeta)
        
        # Botón: Ver últimos reportes
        btn_ultimos = QPushButton("📋 Últimos Reportes")
        btn_ultimos.clicked.connect(self._ver_ultimos_reportes)
        btn_ultimos.setStyleSheet(btn_abrir_carpeta.styleSheet().replace(self.colors['info'], self.colors['success']))
        layout.addWidget(btn_ultimos)
        
        layout.addStretch()
        
        return box
    
    # ==================== MÉTODOS DE GENERACIÓN ====================
    
    def _generar_individual(self):
        """Genera reporte de licitación individual usando ReportWindow."""
        if not REPORT_WINDOW_AVAILABLE:
            QMessageBox.warning(self, "No Disponible", "ReportWindow no está disponible.")
            return
        
        try:
            from app.ui.dialogs.seleccionar_licitacion_dialog import SeleccionarLicitacionDialog
            
            dlg = SeleccionarLicitacionDialog(self.db, self)
            if dlg.exec():
                licitacion = dlg.get_selected_licitacion()
                if licitacion:
                    win = ReportWindow(licitacion, self, start_maximized=True)
                    win.show()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo abrir el reporte:\n{e}")
            import traceback
            traceback.print_exc()
    
    def _abrir_dialogo_reportes(self):
        """Abre el diálogo de KPIs y métricas."""
        if not DIALOGO_REPORTES_AVAILABLE:
            QMessageBox.warning(self, "No Disponible", "DialogoReportes no está disponible.")
            return
        
        try:
            dlg = DialogoReportes(self, self.db)
            dlg.exec()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo abrir el diálogo:\n{e}")
            import traceback
            traceback.print_exc()
    
    def _generar_mensual(self):
        """Genera reporte del mes actual."""
        if not self.reporting:
            QMessageBox.warning(self, "No Disponible", "ReportingEngine no está disponible.")
            return
        
        from PyQt6.QtCore import QDate
        
        try:
            year = QDate.currentDate().year()
            month = QDate.currentDate().month()
            
            reporte = self.reporting.generate_monthly_report(year, month)
            
            # Mostrar resumen
            kpis = reporte['kpis']
            msg = f"""📅 Reporte Mensual Generado
            
📆 Período: {reporte['periodo']}
📍 Desde: {reporte['fecha_inicio']}
📍 Hasta: {reporte['fecha_fin']}

📊 INDICADORES CLAVE:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📋 Total Licitaciones: {kpis['total_licitaciones']}
✅ Adjudicadas: {kpis['licitaciones_adjudicadas']}
🏆 Ganadas: {kpis['licitaciones_ganadas']}
📈 Tasa de Éxito: {kpis['tasa_exito']:.1f}%
💰 Valor Ganado: RD$ {kpis['valor_total_ganado']:,.2f}
💵 Valor Ofertado: RD$ {kpis.get('valor_total_ofertado', 0):,.2f}
"""
            
            QMessageBox.information(self, "Reporte Mensual", msg)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo generar el reporte:\n{e}")
            import traceback
            traceback.print_exc()
    
    def _exportar_excel(self):
        """Exporta todos los datos a Excel con KPIs."""
        if not self.reporting:
            QMessageBox.warning(self, "No Disponible", "ReportingEngine no está disponible.")
            return
        
        from PyQt6.QtCore import QDate
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar a Excel",
            f"reporte_licitaciones_{QDate.currentDate().toString('yyyyMMdd')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            success = self.reporting.export_to_excel(
                filename=filename,
                include_kpis=True
            )
            
            if success:
                resp = QMessageBox.question(
                    self,
                    "Éxito",
                    f"✅ Reporte exportado exitosamente a:\n{filename}\n\n¿Desea abrirlo?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                
                if resp == QMessageBox.StandardButton.Yes:
                    import subprocess
                    subprocess.Popen([filename], shell=True)
            else:
                QMessageBox.warning(self, "Error", "No se pudo exportar el reporte")
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar:\n{e}")
            import traceback
            traceback.print_exc()
    
    def _analisis_financiero(self):
        """Genera análisis financiero detallado con opción de exportar."""
        if not self.db:
            return
        
        try:
            licitaciones = self.db.load_all_licitaciones()
            
            total_base = 0.0
            total_ofertado = 0.0
            total_ganado = 0.0
            ganadas_count = 0
            
            for lic in licitaciones:
                if hasattr(lic, 'get_monto_base_total'):
                    total_base += float(lic.get_monto_base_total() or 0)
                if hasattr(lic, 'get_oferta_total'):
                    total_ofertado += float(lic.get_oferta_total() or 0)
                
                estado = (getattr(lic, 'estado', '') or '').lower()
                if 'ganada' in estado or 'adjudicada' in estado:
                    ganadas_count += 1
                    if hasattr(lic, 'get_oferta_total'):
                        total_ganado += float(lic.get_oferta_total() or 0)
            
            diferencia = total_ofertado - total_base
            efectividad = (total_ganado / total_ofertado * 100) if total_ofertado > 0 else 0
            promedio_ganado = (total_ganado / ganadas_count) if ganadas_count > 0 else 0
            
            msg = f"""💰 ANÁLISIS FINANCIERO
    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    📊 MONTOS TOTALES:
    💵 Monto Base Total: RD$ {total_base:,.2f}
    💼 Monto Ofertado: RD$ {total_ofertado:,.2f}
    🏆 Monto Ganado: RD$ {total_ganado:,.2f}

    📈 INDICADORES:
    📉 Diferencia vs Base: RD$ {diferencia:,.2f} ({diferencia/total_base*100:.1f}%)
    ✅ Efectividad: {efectividad:.1f}%
    📊 Promedio por Licitación Ganada: RD$ {promedio_ganado:,.2f}
    🎯 Total Licitaciones Ganadas: {ganadas_count}
    
    💡 INTERPRETACIÓN:
    {"✅ Ofertas competitivas (bajo presupuesto base)" if diferencia < 0 else "⚠️ Ofertas por encima de la base"}
    {"🎯 Alta efectividad (>50%)" if efectividad > 50 else "⚠️ Efectividad mejorable (<50%)"}

    ¿Desea exportar este análisis a PDF?
    """
            
            from PyQt6.QtWidgets import QMessageBox
            
            resp = QMessageBox.question(
                self,
                "Análisis Financiero",
                msg,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if resp == QMessageBox.StandardButton.Yes:
                self._exportar_analisis_financiero_pdf(
                    total_base, total_ofertado, total_ganado,
                    diferencia, efectividad, promedio_ganado, ganadas_count
                )
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"{e}")
            import traceback
            traceback.print_exc()

    def _exportar_analisis_financiero_pdf(self, total_base, total_ofertado, total_ganado,
                                        diferencia, efectividad, promedio_ganado, ganadas_count):
        """Exporta el análisis financiero a PDF."""
        from PyQt6.QtCore import QDate
        from datetime import datetime
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar Análisis Financiero",
            f"analisis_financiero_{QDate.currentDate().toString('yyyyMMdd')}.pdf",
            "PDF Files (*.pdf)"
        )
        
        if not filename:
            return
        
        try:
            # Intentar usar ReportLab si está disponible
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
                from reportlab.lib import colors
                
                doc = SimpleDocTemplate(filename, pagesize=letter)
                elements = []
                styles = getSampleStyleSheet()
                
                # Título
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=24,
                    textColor=colors.HexColor('#7C4DFF'),
                    spaceAfter=30,
                    alignment=1
                )
                elements.append(Paragraph("💰 ANÁLISIS FINANCIERO", title_style))
                elements.append(Spacer(1, 0.3*inch))
                
                # Fecha
                elements.append(Paragraph(
                    f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                    styles['Normal']
                ))
                elements.append(Spacer(1, 0.3*inch))
                
                # Tabla de montos
                data_montos = [
                    ['Concepto', 'Monto (RD$)'],
                    ['Monto Base Total', f'{total_base:,.2f}'],
                    ['Monto Ofertado', f'{total_ofertado:,.2f}'],
                    ['Monto Ganado', f'{total_ganado:,.2f}'],
                    ['Diferencia vs Base', f'{diferencia:,.2f}'],
                ]
                
                table_montos = Table(data_montos, colWidths=[3*inch, 2*inch])
                table_montos.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7C4DFF')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                
                elements.append(table_montos)
                elements.append(Spacer(1, 0.5*inch))
                
                # Tabla de indicadores
                data_indicadores = [
                    ['Indicador', 'Valor'],
                    ['Efectividad', f'{efectividad:.1f}%'],
                    ['Promedio por Ganada', f'RD$ {promedio_ganado:,.2f}'],
                    ['Total Ganadas', str(ganadas_count)],
                    ['% vs Base', f'{diferencia/total_base*100:.1f}%' if total_base > 0 else 'N/A'],
                ]
                
                table_indicadores = Table(data_indicadores, colWidths=[3*inch, 2*inch])
                table_indicadores.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00C853')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                
                elements.append(table_indicadores)
                elements.append(Spacer(1, 0.5*inch))
                
                # Interpretación
                interpretacion = []
                if diferencia < 0:
                    interpretacion.append("✅ Las ofertas están bajo el presupuesto base (competitivas)")
                else:
                    interpretacion.append("⚠️ Las ofertas están por encima del presupuesto base")
                
                if efectividad > 50:
                    interpretacion.append("🎯 Alta efectividad en licitaciones (>50%)")
                else:
                    interpretacion.append("⚠️ La efectividad es mejorable (<50%)")
                
                elements.append(Paragraph("<b>💡 INTERPRETACIÓN:</b>", styles['Heading2']))
                for interp in interpretacion:
                    elements.append(Paragraph(f"• {interp}", styles['Normal']))
                
                # Generar PDF
                doc.build(elements)
                
                QMessageBox.information(
                    self,
                    "Éxito",
                    f"✅ Análisis financiero exportado a:\n{filename}"
                )
                
                # Preguntar si quiere abrirlo
                resp = QMessageBox.question(
                    self,
                    "Abrir PDF",
                    "¿Desea abrir el PDF generado?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                
                if resp == QMessageBox.StandardButton.Yes:
                    import subprocess
                    subprocess.Popen([filename], shell=True)
            
            except ImportError:
                QMessageBox.warning(
                    self,
                    "ReportLab no disponible",
                    "Para exportar a PDF, instala ReportLab:\n\npip install reportlab"
                )
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo exportar:\n{e}")
            import traceback
            traceback.print_exc()
    
    def _reporte_ganadas(self):
        """Genera reporte de licitaciones ganadas con opción de exportar."""
        if not self.db:
            return
        
        try:
            licitaciones = self.db.load_all_licitaciones()
            ganadas = [lic for lic in licitaciones 
                    if 'ganada' in (getattr(lic, 'estado', '') or '').lower() 
                    or 'adjudicada' in (getattr(lic, 'estado', '') or '').lower()]
            
            if not ganadas:
                QMessageBox.information(self, "Sin Datos", "No hay licitaciones ganadas en el sistema.")
                return
            
            total_ganado = sum(
                float(lic.get_oferta_total() or 0) 
                for lic in ganadas 
                if hasattr(lic, 'get_oferta_total')
            )
            
            msg = f"""🏆 LICITACIONES GANADAS
    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    📊 RESUMEN:
    ✅ Total Ganadas: {len(ganadas)}
    💰 Valor Total: RD$ {total_ganado:,.2f}
    📈 Promedio: RD$ {total_ganado / len(ganadas):,.2f}

    📋 PRIMERAS 10:
    """
            
            for i, lic in enumerate(ganadas[:10], 1):
                codigo = getattr(lic, 'numero_proceso', 'N/A')
                nombre = getattr(lic, 'nombre_proceso', 'N/A')[:45]
                monto = float(lic.get_oferta_total() or 0) if hasattr(lic, 'get_oferta_total') else 0
                msg += f"\n{i}. {codigo}\n   {nombre}...\n   RD$ {monto:,.2f}"
            
            if len(ganadas) > 10:
                msg += f"\n\n... y {len(ganadas) - 10} más"
            
            msg += "\n\n¿Desea exportar a Excel?"
            
            resp = QMessageBox.question(
                self,
                "Licitaciones Ganadas",
                msg,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if resp == QMessageBox.StandardButton.Yes:
                self._exportar_ganadas_excel(ganadas, total_ganado)
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"{e}")

    def _exportar_ganadas_excel(self, ganadas, total_ganado):
        """Exporta licitaciones ganadas a Excel."""
        from PyQt6.QtCore import QDate
        
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar Licitaciones Ganadas",
            f"licitaciones_ganadas_{QDate.currentDate().toString('yyyyMMdd')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Licitaciones Ganadas"
            
            # Encabezados
            headers = ["#", "Código", "Nombre", "Institución", "Estado", "Monto Ofertado"]
            ws.append(headers)
            
            # Estilo de encabezados
            header_fill = PatternFill(start_color="7C4DFF", end_color="7C4DFF", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Datos
            for i, lic in enumerate(ganadas, 1):
                codigo = getattr(lic, 'numero_proceso', 'N/A')
                nombre = getattr(lic, 'nombre_proceso', 'N/A')
                institucion = getattr(lic, 'institucion', 'N/A')
                estado = getattr(lic, 'estado', 'N/A')
                monto = float(lic.get_oferta_total() or 0) if hasattr(lic, 'get_oferta_total') else 0
                
                ws.append([i, codigo, nombre, institucion, estado, monto])
            
            # Fila de totales
            ws.append([])
            total_row = ws.max_row + 1
            ws.cell(total_row, 1, "TOTAL")
            ws.cell(total_row, 5, len(ganadas))
            ws.cell(total_row, 6, total_ganado)
            
            # Estilo de totales
            for col in [1, 5, 6]:
                cell = ws.cell(total_row, col)
                cell.font = Font(bold=True)
            
            # Ajustar anchos
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 18
            
            # Guardar
            wb.save(filename)
            
            QMessageBox.information(
                self,
                "Éxito",
                f"✅ Listado exportado a:\n{filename}"
            )
            
            # Abrir
            resp = QMessageBox.question(
                self,
                "Abrir Excel",
                "¿Desea abrir el archivo?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if resp == QMessageBox.StandardButton.Yes:
                import subprocess
                subprocess.Popen([filename], shell=True)
        
        except ImportError:
            QMessageBox.warning(
                self,
                "OpenPyXL no disponible",
                "Para exportar a Excel, instala OpenPyXL:\n\npip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo exportar:\n{e}")
    
    def _abrir_carpeta_reportes(self):
        """Abre la carpeta donde se guardan los reportes."""
        import subprocess
        import sys
        
        carpeta = os.path.expanduser("~/Documents/Reportes_Licitaciones")
        os.makedirs(carpeta, exist_ok=True)
        
        if sys.platform == 'win32':
            subprocess.Popen(['explorer', carpeta])
        elif sys.platform == 'darwin':
            subprocess.Popen(['open', carpeta])
        else:
            subprocess.Popen(['xdg-open', carpeta])
    
    def _ver_ultimos_reportes(self):
        """Muestra lista de últimos reportes generados."""
        QMessageBox.information(
            self,
            "Últimos Reportes",
            "Funcionalidad en desarrollo.\n\n"
            "Mostrará historial de reportes generados con opciones para:\n"
            "• Ver reporte\n"
            "• Reabrir en aplicación\n"
            "• Eliminar\n"
            "• Regenerar"
        )