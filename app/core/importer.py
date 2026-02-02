"""
Importer Module - Importadores y conectores

Este módulo maneja la importación de datos desde Excel/CSV, con validación
y preview de datos antes de importar.

Soporta importación de:
- Lotes
- Documentos checklist
- Competidores
- Tareas
- Licitaciones completas
"""
from __future__ import annotations

import csv
from typing import Any, Dict, List, Optional, Tuple, Callable
from pathlib import Path
from dataclasses import dataclass, field

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


@dataclass
class ImportResult:
    """Resultado de una operación de importación."""
    success: bool = False
    total_rows: int = 0
    valid_rows: int = 0
    invalid_rows: int = 0
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    imported_data: List[Dict[str, Any]] = field(default_factory=list)

    def add_error(self, row: int, message: str):
        """Añade un error a la lista."""
        self.errors.append(f"Fila {row}: {message}")
        self.invalid_rows += 1

    def add_warning(self, row: int, message: str):
        """Añade una advertencia a la lista."""
        self.warnings.append(f"Fila {row}: {message}")


class ExcelImporter:
    """Importador de datos desde archivos Excel/CSV."""

    def __init__(self):
        self.column_mappings = {
            "lotes": {
                "numero": ["numero", "número", "num", "no"],
                "nombre": ["nombre", "descripcion", "descripción"],
                "monto_base": ["monto_base", "monto base", "presupuesto", "valor"],
                "monto_ofertado": ["monto_ofertado", "oferta", "monto ofertado"],
            },
            "documentos": {
                "codigo": ["codigo", "código", "cod"],
                "nombre": ["nombre", "documento", "descripcion"],
                "categoria": ["categoria", "categoría", "tipo"],
                "obligatorio": ["obligatorio", "requerido", "required"],
                "subsanable": ["subsanable", "sub"],
            },
            "competidores": {
                "nombre": ["nombre", "empresa", "razon_social"],
                "rnc": ["rnc", "ruc", "identificacion"],
                "categoria": ["categoria", "categoría", "rubro"],
            },
            "tareas": {
                "titulo": ["titulo", "título", "tarea", "task"],
                "descripcion": ["descripcion", "descripción", "detalle"],
                "responsable": ["responsable", "asignado", "assigned"],
                "fecha_limite": ["fecha_limite", "deadline", "vencimiento"],
                "prioridad": ["prioridad", "priority"],
            }
        }

    def read_excel(self, file_path: str) -> Tuple[List[str], List[List[Any]]]:
        """
        Lee un archivo Excel y retorna encabezados y filas.

        Args:
            file_path: Ruta del archivo Excel

        Returns:
            Tupla con (encabezados, filas)
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl no está instalado. Instálalo con: pip install openpyxl")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # Leer encabezados (primera fila)
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value or "").strip())

        # Leer filas
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Saltar filas vacías
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            rows.append(list(row))

        return headers, rows

    def read_csv(self, file_path: str, delimiter: str = ",") -> Tuple[List[str], List[List[Any]]]:
        """
        Lee un archivo CSV y retorna encabezados y filas.

        Args:
            file_path: Ruta del archivo CSV
            delimiter: Delimitador usado (por defecto coma)

        Returns:
            Tupla con (encabezados, filas)
        """
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f, delimiter=delimiter)
            headers = next(reader)
            rows = [row for row in reader if any(cell.strip() for cell in row)]

        return headers, rows

    def map_columns(
        self,
        headers: List[str],
        entity_type: str
    ) -> Dict[str, int]:
        """
        Mapea automáticamente las columnas del archivo a los campos esperados.

        Args:
            headers: Lista de encabezados del archivo
            entity_type: Tipo de entidad (lotes, documentos, competidores, tareas)

        Returns:
            Diccionario {campo_esperado: índice_columna}
        """
        if entity_type not in self.column_mappings:
            raise ValueError(f"Tipo de entidad no soportado: {entity_type}")

        mapping = {}
        field_mappings = self.column_mappings[entity_type]

        # Normalizar encabezados
        normalized_headers = [h.lower().strip() for h in headers]

        for field, possible_names in field_mappings.items():
            for idx, header in enumerate(normalized_headers):
                if header in possible_names:
                    mapping[field] = idx
                    break

        return mapping

    def validate_row(
        self,
        row: List[Any],
        column_mapping: Dict[str, int],
        required_fields: List[str]
    ) -> Tuple[bool, Optional[str]]:
        """
        Valida una fila de datos.

        Args:
            row: Datos de la fila
            column_mapping: Mapeo de columnas
            required_fields: Campos requeridos

        Returns:
            Tupla (es_valida, mensaje_error)
        """
        for field in required_fields:
            if field not in column_mapping:
                return False, f"Campo requerido '{field}' no encontrado en el archivo"
            
            idx = column_mapping[field]
            if idx >= len(row):
                return False, f"Fila incompleta, falta campo '{field}'"
            
            value = row[idx]
            if value is None or str(value).strip() == "":
                return False, f"Campo requerido '{field}' está vacío"

        return True, None

    def import_lotes(
        self,
        file_path: str,
        licitacion_id: str,
        db_adapter
    ) -> ImportResult:
        """
        Importa lotes desde un archivo Excel/CSV.

        Args:
            file_path: Ruta del archivo
            licitacion_id: ID de la licitación a la que agregar los lotes
            db_adapter: Adaptador de base de datos

        Returns:
            Objeto ImportResult con el resultado de la importación
        """
        result = ImportResult()

        try:
            # Leer archivo
            if file_path.endswith('.csv'):
                headers, rows = self.read_csv(file_path)
            else:
                headers, rows = self.read_excel(file_path)

            result.total_rows = len(rows)

            # Mapear columnas
            mapping = self.map_columns(headers, "lotes")
            required_fields = ["numero", "nombre"]

            # Validar y procesar cada fila
            lotes_data = []
            for idx, row in enumerate(rows, start=2):  # Empezar en 2 (fila 1 es header)
                is_valid, error_msg = self.validate_row(row, mapping, required_fields)
                
                if not is_valid:
                    result.add_error(idx, error_msg)
                    continue

                # Extraer datos
                lote_data = {
                    "numero": str(row[mapping["numero"]]).strip(),
                    "nombre": str(row[mapping["nombre"]]).strip(),
                    "monto_base": self._parse_float(row, mapping.get("monto_base"), 0.0),
                    "monto_ofertado": self._parse_float(row, mapping.get("monto_ofertado"), 0.0),
                    "participamos": True,
                }

                lotes_data.append(lote_data)
                result.valid_rows += 1

            # Importar lotes válidos
            if lotes_data:
                # Cargar licitación
                licitacion = db_adapter.load_licitacion_by_id(licitacion_id)
                if not licitacion:
                    result.add_error(0, f"Licitación {licitacion_id} no encontrada")
                    return result

                # Agregar lotes
                from app.core.models import Lote
                for lote_data in lotes_data:
                    lote = Lote(**lote_data)
                    licitacion.lotes.append(lote)

                # Guardar cambios
                db_adapter.save_licitacion(licitacion)
                result.imported_data = lotes_data
                result.success = True

        except Exception as e:
            result.add_error(0, f"Error general: {str(e)}")

        return result

    def import_documentos(
        self,
        file_path: str,
        licitacion_id: str,
        db_adapter
    ) -> ImportResult:
        """
        Importa documentos checklist desde un archivo Excel/CSV.

        Args:
            file_path: Ruta del archivo
            licitacion_id: ID de la licitación
            db_adapter: Adaptador de base de datos

        Returns:
            Objeto ImportResult con el resultado de la importación
        """
        result = ImportResult()

        try:
            # Leer archivo
            if file_path.endswith('.csv'):
                headers, rows = self.read_csv(file_path)
            else:
                headers, rows = self.read_excel(file_path)

            result.total_rows = len(rows)

            # Mapear columnas
            mapping = self.map_columns(headers, "documentos")
            required_fields = ["codigo", "nombre"]

            # Validar y procesar cada fila
            docs_data = []
            for idx, row in enumerate(rows, start=2):
                is_valid, error_msg = self.validate_row(row, mapping, required_fields)
                
                if not is_valid:
                    result.add_error(idx, error_msg)
                    continue

                # Extraer datos
                doc_data = {
                    "codigo": str(row[mapping["codigo"]]).strip(),
                    "nombre": str(row[mapping["nombre"]]).strip(),
                    "categoria": self._get_value(row, mapping.get("categoria"), "General"),
                    "obligatorio": self._parse_bool(row, mapping.get("obligatorio"), False),
                    "subsanable": self._get_value(row, mapping.get("subsanable"), "Subsanable"),
                }

                docs_data.append(doc_data)
                result.valid_rows += 1

            # Importar documentos válidos
            if docs_data:
                licitacion = db_adapter.load_licitacion_by_id(licitacion_id)
                if not licitacion:
                    result.add_error(0, f"Licitación {licitacion_id} no encontrada")
                    return result

                from app.core.models import Documento
                for doc_data in docs_data:
                    doc = Documento(**doc_data)
                    licitacion.documentos_solicitados.append(doc)

                db_adapter.save_licitacion(licitacion)
                result.imported_data = docs_data
                result.success = True

        except Exception as e:
            result.add_error(0, f"Error general: {str(e)}")

        return result

    def _parse_float(self, row: List[Any], idx: Optional[int], default: float = 0.0) -> float:
        """Parsea un valor como float."""
        if idx is None or idx >= len(row):
            return default
        try:
            value = row[idx]
            if value is None or str(value).strip() == "":
                return default
            return float(str(value).replace(",", ""))
        except (ValueError, TypeError):
            return default

    def _parse_bool(self, row: List[Any], idx: Optional[int], default: bool = False) -> bool:
        """Parsea un valor como booleano."""
        if idx is None or idx >= len(row):
            return default
        value = str(row[idx]).strip().lower()
        return value in ["true", "yes", "sí", "si", "1", "x", "✓"]

    def _get_value(self, row: List[Any], idx: Optional[int], default: str = "") -> str:
        """Obtiene un valor como string."""
        if idx is None or idx >= len(row):
            return default
        value = row[idx]
        if value is None:
            return default
        return str(value).strip()

    def preview_import(
        self,
        file_path: str,
        entity_type: str,
        max_rows: int = 10
    ) -> Dict[str, Any]:
        """
        Genera un preview de los datos a importar.

        Args:
            file_path: Ruta del archivo
            entity_type: Tipo de entidad
            max_rows: Número máximo de filas a previsualizar

        Returns:
            Diccionario con información del preview
        """
        try:
            if file_path.endswith('.csv'):
                headers, rows = self.read_csv(file_path)
            else:
                headers, rows = self.read_excel(file_path)

            mapping = self.map_columns(headers, entity_type)
            
            preview_rows = rows[:max_rows]

            return {
                "success": True,
                "total_rows": len(rows),
                "headers": headers,
                "column_mapping": mapping,
                "preview_rows": preview_rows,
                "entity_type": entity_type,
            }

        except Exception as e:
            return {
                "success": False,
                "error": str(e),
            }
