"""
Reporting Module - Reportería avanzada y KPIs

Este módulo genera reportes avanzados con KPIs, exportación a Excel/PDF,
y análisis estadístico de licitaciones.

KPIs principales:
- Tasa de adjudicación
- Tiempo de ciclo promedio
- Causas de pérdida
- % de completitud de documentos
- Vencimientos próximos
"""
from __future__ import annotations

import datetime
import statistics
from typing import Any, Dict, List, Optional, Tuple
from dataclasses import dataclass, field

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


@dataclass
class ReportKPIs:
    """Contenedor de KPIs calculados."""
    total_licitaciones: int = 0
    licitaciones_adjudicadas: int = 0
    licitaciones_ganadas: int = 0
    tasa_adjudicacion: float = 0.0
    tasa_exito: float = 0.0
    tiempo_ciclo_promedio: float = 0.0
    valor_total_ofertado: float = 0.0
    valor_total_ganado: float = 0.0
    completitud_documentos_promedio: float = 0.0
    vencimientos_proximos: int = 0
    causas_perdida: Dict[str, int] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        """Convierte los KPIs a diccionario."""
        return {
            "total_licitaciones": self.total_licitaciones,
            "licitaciones_adjudicadas": self.licitaciones_adjudicadas,
            "licitaciones_ganadas": self.licitaciones_ganadas,
            "tasa_adjudicacion": round(self.tasa_adjudicacion, 2),
            "tasa_exito": round(self.tasa_exito, 2),
            "tiempo_ciclo_promedio": round(self.tiempo_ciclo_promedio, 2),
            "valor_total_ofertado": round(self.valor_total_ofertado, 2),
            "valor_total_ganado": round(self.valor_total_ganado, 2),
            "completitud_documentos_promedio": round(self.completitud_documentos_promedio, 2),
            "vencimientos_proximos": self.vencimientos_proximos,
            "causas_perdida": self.causas_perdida,
        }


class ReportingEngine:
    """Motor de generación de reportes y análisis."""

    def __init__(self, db_adapter):
        """
        Inicializa el motor de reportes.

        Args:
            db_adapter: Adaptador de base de datos para acceder a los datos
        """
        self.db = db_adapter

    def calculate_kpis(
        self,
        fecha_inicio: Optional[str] = None,
        fecha_fin: Optional[str] = None,
        institucion: Optional[str] = None
    ) -> ReportKPIs:
        """
        Calcula los KPIs para un período y filtros específicos.

        Args:
            fecha_inicio: Fecha de inicio del período (ISO format)
            fecha_fin: Fecha fin del período (ISO format)
            institucion: Filtrar por institución específica

        Returns:
            Objeto ReportKPIs con los indicadores calculados
        """
        # Obtener todas las licitaciones
        licitaciones = self.db.load_all_licitaciones()

        # Aplicar filtros
        if fecha_inicio:
            licitaciones = [
                lic for lic in licitaciones
                if hasattr(lic, 'fecha_creacion') and str(lic.fecha_creacion) >= fecha_inicio
            ]
        if fecha_fin:
            licitaciones = [
                lic for lic in licitaciones
                if hasattr(lic, 'fecha_creacion') and str(lic.fecha_creacion) <= fecha_fin
            ]
        if institucion:
            licitaciones = [
                lic for lic in licitaciones
                if lic.institucion == institucion
            ]

        kpis = ReportKPIs()
        kpis.total_licitaciones = len(licitaciones)

        if not licitaciones:
            return kpis

        # Calcular métricas
        adjudicadas = [lic for lic in licitaciones if lic.adjudicada]
        kpis.licitaciones_adjudicadas = len(adjudicadas)

        # Verificar si alguna de nuestras empresas ganó
        kpis.licitaciones_ganadas = 0
        for lic in adjudicadas:
            for empresa in lic.empresas_nuestras:
                if empresa.nombre in lic.adjudicada_a:
                    kpis.licitaciones_ganadas += 1
                    break

        # Tasas
        if kpis.total_licitaciones > 0:
            kpis.tasa_adjudicacion = (kpis.licitaciones_adjudicadas / kpis.total_licitaciones) * 100
        if kpis.licitaciones_adjudicadas > 0:
            kpis.tasa_exito = (kpis.licitaciones_ganadas / kpis.licitaciones_adjudicadas) * 100

        # Valor total ofertado y ganado
        for lic in licitaciones:
            for lote in lic.lotes:
                if lote.participamos:
                    kpis.valor_total_ofertado += lote.monto_ofertado
                if lote.ganado_por_nosotros:
                    kpis.valor_total_ganado += lote.monto_ofertado

        # Completitud de documentos
        completitudes = []
        for lic in licitaciones:
            if lic.documentos_solicitados:
                presentados = sum(1 for doc in lic.documentos_solicitados if doc.presentado)
                total = len(lic.documentos_solicitados)
                completitud = (presentados / total) * 100 if total > 0 else 0
                completitudes.append(completitud)
        
        if completitudes:
            kpis.completitud_documentos_promedio = statistics.mean(completitudes)

        # Causas de pérdida
        perdidas = [lic for lic in licitaciones if lic.adjudicada and not any(
            empresa.nombre in lic.adjudicada_a for empresa in lic.empresas_nuestras
        )]
        
        for lic in perdidas:
            motivo = lic.motivo_descalificacion or "No especificado"
            kpis.causas_perdida[motivo] = kpis.causas_perdida.get(motivo, 0) + 1

        # Vencimientos próximos (próximos 7 días)
        now = datetime.datetime.now().date()
        limite = now + datetime.timedelta(days=7)
        
        for lic in licitaciones:
            if not lic.adjudicada:  # Solo licitaciones activas
                for fase, datos in lic.cronograma.items():
                    fecha_str = datos.get('fecha', '')
                    if fecha_str:
                        try:
                            fecha = datetime.datetime.strptime(fecha_str, "%Y-%m-%d").date()
                            if now <= fecha <= limite:
                                kpis.vencimientos_proximos += 1
                        except (ValueError, TypeError):
                            pass

        return kpis

    def export_to_excel(
        self,
        filename: str,
        licitaciones: Optional[List] = None,
        include_kpis: bool = True
    ) -> bool:
        """
        Exporta datos a un archivo Excel.

        Args:
            filename: Ruta del archivo a crear
            licitaciones: Lista de licitaciones a exportar (None = todas)
            include_kpis: Si incluir una hoja con KPIs

        Returns:
            True si se exportó exitosamente
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl no está instalado. Instálalo con: pip install openpyxl")

        if licitaciones is None:
            licitaciones = self.db.load_all_licitaciones()

        wb = openpyxl.Workbook()
        
        # Hoja de licitaciones
        ws = wb.active
        ws.title = "Licitaciones"
        
        # Encabezados
        headers = [
            "ID", "Número Proceso", "Nombre", "Institución", "Estado",
            "Fase A", "Fase B", "Adjudicada", "Adjudicada A",
            "Fecha Creación", "Valor Ofertado", "Valor Ganado"
        ]
        ws.append(headers)
        
        # Estilo de encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Datos
        for lic in licitaciones:
            valor_ofertado = sum(lote.monto_ofertado for lote in lic.lotes if lote.participamos)
            valor_ganado = sum(lote.monto_ofertado for lote in lic.lotes if lote.ganado_por_nosotros)
            
            row = [
                lic.id or "",
                lic.numero_proceso,
                lic.nombre_proceso,
                lic.institucion,
                lic.estado,
                "Sí" if lic.fase_A_superada else "No",
                "Sí" if lic.fase_B_superada else "No",
                "Sí" if lic.adjudicada else "No",
                lic.adjudicada_a or "",
                str(lic.fecha_creacion) if hasattr(lic, 'fecha_creacion') else "",
                valor_ofertado,
                valor_ganado,
            ]
            ws.append(row)

        # Ajustar anchos de columna
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Hoja de KPIs
        if include_kpis:
            kpis = self.calculate_kpis()
            ws_kpis = wb.create_sheet("KPIs")
            
            ws_kpis.append(["Indicador", "Valor"])
            ws_kpis.append(["Total Licitaciones", kpis.total_licitaciones])
            ws_kpis.append(["Licitaciones Adjudicadas", kpis.licitaciones_adjudicadas])
            ws_kpis.append(["Licitaciones Ganadas", kpis.licitaciones_ganadas])
            ws_kpis.append(["Tasa de Adjudicación (%)", f"{kpis.tasa_adjudicacion:.2f}"])
            ws_kpis.append(["Tasa de Éxito (%)", f"{kpis.tasa_exito:.2f}"])
            ws_kpis.append(["Valor Total Ofertado", f"{kpis.valor_total_ofertado:,.2f}"])
            ws_kpis.append(["Valor Total Ganado", f"{kpis.valor_total_ganado:,.2f}"])
            ws_kpis.append(["Completitud Documentos (%)", f"{kpis.completitud_documentos_promedio:.2f}"])
            ws_kpis.append(["Vencimientos Próximos (7 días)", kpis.vencimientos_proximos])
            
            # Estilo
            for cell in ws_kpis[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            ws_kpis.column_dimensions['A'].width = 30
            ws_kpis.column_dimensions['B'].width = 20

        # Guardar
        try:
            wb.save(filename)
            return True
        except Exception as e:
            print(f"Error al guardar Excel: {e}")
            return False

    def generate_monthly_report(
        self,
        year: int,
        month: int
    ) -> Dict[str, Any]:
        """
        Genera un reporte mensual completo.

        Args:
            year: Año del reporte
            month: Mes del reporte (1-12)

        Returns:
            Diccionario con datos del reporte mensual
        """
        # Calcular fechas de inicio y fin del mes
        fecha_inicio = datetime.date(year, month, 1)
        if month == 12:
            fecha_fin = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            fecha_fin = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)

        kpis = self.calculate_kpis(
            fecha_inicio=fecha_inicio.isoformat(),
            fecha_fin=fecha_fin.isoformat()
        )

        return {
            "periodo": f"{year}-{month:02d}",
            "fecha_inicio": fecha_inicio.isoformat(),
            "fecha_fin": fecha_fin.isoformat(),
            "kpis": kpis.to_dict(),
            "generado_en": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        }

    def get_licitaciones_by_estado(self, estado: str) -> List:
        """Obtiene licitaciones filtradas por estado."""
        licitaciones = self.db.load_all_licitaciones()
        return [lic for lic in licitaciones if lic.estado == estado]

    def get_licitaciones_by_institucion(self, institucion: str) -> List:
        """Obtiene licitaciones filtradas por institución."""
        licitaciones = self.db.load_all_licitaciones()
        return [lic for lic in licitaciones if lic.institucion == institucion]
